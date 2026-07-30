import math
from io import BytesIO
from pathlib import Path
from datetime import datetime

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils.units import pixels_to_EMU
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportlabImage

from database.conexion import get_connection


def exportar_bitacora_excel():
    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        if conexion is None:
            raise HTTPException(status_code=500, detail="No fue posible conectar a la base de datos")

        cursor = conexion.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT
                b.id_bitacora AS id,
                b.fecha,
                COALESCE(u.nombre, u.usuario, 'Sistema') AS usuario,
                b.accion,
                b.tabla_afectada AS modulo,
                b.detalle
            FROM tbl_bitacora b
            LEFT JOIN tbl_usuario u ON u.id_usuario = b.id_usuario
            ORDER BY b.fecha DESC
            """
        )
        datos = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Bitácora"
        ws.sheet_view.showGridLines = True

        azul = "0A4B8F"
        blanco = "FFFFFF"

        ws.column_dimensions["A"].width = 20

        ws.row_dimensions[1].height = 12
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 24
        ws.row_dimensions[5].height = 12

        ws.merge_cells("A1:A4")
        fill_blanco = PatternFill("solid", fgColor=blanco)
        for row in ws["A1:A4"]:
            for cell in row:
                cell.fill = fill_blanco

        logo = Path(__file__).parent.parent / "assets" / "hospital_logo.png"
        if logo.exists():
            imagen = Image(str(logo))
            imagen.width = 120
            imagen.height = 120
            marker = AnchorMarker(col=0, row=0, colOff=pixels_to_EMU(35), rowOff=pixels_to_EMU(8))
            imagen.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(imagen.width), pixels_to_EMU(imagen.height)))
            ws.add_image(imagen)

        fill_azul = PatternFill("solid", fgColor=azul)
        for fila in range(2, 5):
            for columna in range(2, 7):
                ws.cell(row=fila, column=columna).fill = fill_azul

        ws.merge_cells("B2:F2")
        ws["B2"] = "HOSPITAL MILITAR"
        ws["B2"].font = Font(bold=True, size=16, color=blanco)
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B3:F3")
        ws["B3"] = "SISTEMA SIGIP"
        ws["B3"].font = Font(bold=True, size=12, color=blanco)
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B4:F4")
        ws["B4"] = "REPORTE DE BITÁCORA"
        ws["B4"].font = Font(bold=True, size=11, color=blanco)
        ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

        fecha = datetime.now()
        ws["B6"] = "Fecha"
        ws["C6"] = fecha.strftime("%d/%m/%Y")
        ws["B7"] = "Hora"
        ws["C7"] = fecha.strftime("%H:%M")
        ws["E6"] = "Total Registros"
        ws["F6"] = len(datos)
        ws["E7"] = "Registros"
        ws["F7"] = len(datos)

        for celda in ["B6", "B7", "E6", "E7"]:
            ws[celda].font = Font(bold=True)

        ws["B6"].alignment = Alignment(horizontal="left")
        ws["B7"].alignment = Alignment(horizontal="left")
        ws["C6"].alignment = Alignment(horizontal="left")
        ws["C7"].alignment = Alignment(horizontal="left")
        ws["E6"].alignment = Alignment(horizontal="left")
        ws["E7"].alignment = Alignment(horizontal="left")
        ws["F6"].alignment = Alignment(horizontal="right")
        ws["F7"].alignment = Alignment(horizontal="right")

        encabezados = ["Fecha / Hora", "Usuario", "Acción", "Módulo", "Detalle"]
        fila_inicio = 10
        borde = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))

        for columna, texto in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_inicio, column=columna)
            celda.value = texto
            celda.font = Font(bold=True, color=blanco)
            celda.fill = PatternFill(fill_type="solid", fgColor=azul)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            celda.border = borde

        fila = fila_inicio + 1
        for item in datos:
            ws.cell(fila, 1).value = item.get("fecha", "") or ""
            ws.cell(fila, 2).value = item.get("usuario", "") or ""
            ws.cell(fila, 3).value = item.get("accion", "") or ""
            ws.cell(fila, 4).value = item.get("modulo", "") or ""
            ws.cell(fila, 5).value = item.get("detalle", "") or ""
            for columna in range(1, 6):
                celda = ws.cell(fila, columna)
                celda.border = borde
                celda.alignment = Alignment(vertical="center", wrapText=True)
            row_values = [str(ws.cell(fila, col).value or "") for col in range(1, 6)]
            max_len = max(len(text) for text in row_values) if row_values else 0
            lines = max(1, math.ceil(max_len / 40))
            ws.row_dimensions[fila].height = min(45, 16 + lines * 10)
            fila += 1

        ultima_fila = max(fila - 1, fila_inicio)
        if datos:
            tabla = ExcelTable(displayName="TablaBitacora", ref=f"A{fila_inicio}:E{ultima_fila}")
            tabla.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            ws.add_table(tabla)

        ws.freeze_panes = f"A{fila_inicio + 1}"
        for columna in range(1, 6):
            letra = get_column_letter(columna)
            longitud = 0
            for fila_excel in range(1, ultima_fila + 1):
                valor = ws.cell(fila_excel, columna).value
                if valor is not None:
                    longitud = max(longitud, len(str(valor)))
            ws.column_dimensions[letra].width = max(longitud + 4, 15)
        ws.column_dimensions["E"].width = 55

        fila_resumen = ultima_fila + 3
        ws.merge_cells(start_row=fila_resumen, start_column=1, end_row=fila_resumen, end_column=5)
        celda_resumen = ws.cell(row=fila_resumen, column=1)
        celda_resumen.value = "Reporte generado automáticamente por SIGIP"
        celda_resumen.font = Font(italic=True, size=10, color="666666")
        celda_resumen.alignment = Alignment(horizontal="center")

        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

        wb.properties.creator = "SIGIP"
        wb.properties.title = "Reporte de Bitácora"
        wb.properties.subject = "Hospital Militar"
        wb.properties.company = "Hospital Militar"
        wb.properties.category = "Reportes"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar bitácora a Excel: {str(e)}")
    finally:
        if conexion and conexion.is_connected() and cursor is not None:
            cursor.close()
            conexion.close()


def exportar_bitacora_pdf():
    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        if conexion is None:
            raise HTTPException(status_code=500, detail="No fue posible conectar a la base de datos")

        cursor = conexion.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT
                b.id_bitacora AS id,
                b.fecha,
                COALESCE(u.nombre, u.usuario, 'Sistema') AS usuario,
                b.accion,
                b.tabla_afectada AS modulo,
                b.detalle
            FROM tbl_bitacora b
            LEFT JOIN tbl_usuario u ON u.id_usuario = b.id_usuario
            ORDER BY b.fecha DESC
            """
        )
        datos = cursor.fetchall()

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=1.2 * cm,
            leftMargin=1.2 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#003366"),
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#555555"),
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9,
            leading=11,
            alignment=TA_LEFT,
        )

        story = []
        logo_path = Path(__file__).parent.parent / "assets" / "hospital_logo.png"
        if logo_path.exists():
            try:
                logo = ReportlabImage(str(logo_path), width=2.3 * cm, height=2.3 * cm)
                header_table = Table([[logo, Paragraph("<b>HOSPITAL MILITAR</b><br/>SIGIP - Sistema de Gestión de Direcciones IP<br/><font color='#003366'><b>REPORTE DE BITÁCORA</b></font>", title_style)]], colWidths=[2.8 * cm, 22.1 * cm])
                header_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('ALIGN', (0, 0), (0, 0), 'LEFT'), ('ALIGN', (1, 0), (1, 0), 'CENTER')]))
                story.append(header_table)
            except Exception:
                story.append(Paragraph("HOSPITAL MILITAR\nSIGIP - Sistema de Gestión de Direcciones IP\nREPORTE DE BITÁCORA", title_style))
        else:
            story.append(Paragraph("HOSPITAL MILITAR\nSIGIP - Sistema de Gestión de Direcciones IP\nREPORTE DE BITÁCORA", title_style))
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total registros: {len(datos)}", subtitle_style))
        story.append(Spacer(1, 0.3 * cm))

        table_data = [[Paragraph("Fecha / Hora", body_style), Paragraph("Usuario", body_style), Paragraph("Acción", body_style), Paragraph("Módulo", body_style), Paragraph("Detalle", body_style)]]
        for item in datos:
            table_data.append([
                Paragraph(str(item.get("fecha", "") or ""), body_style),
                Paragraph(str(item.get("usuario", "") or ""), body_style),
                Paragraph(str(item.get("accion", "") or ""), body_style),
                Paragraph(str(item.get("modulo", "") or ""), body_style),
                Paragraph(str(item.get("detalle", "") or ""), body_style),
            ])

        table = Table(table_data, repeatRows=1, hAlign="LEFT", colWidths=[3.3 * cm, 3.2 * cm, 2.8 * cm, 3.2 * cm, 10.2 * cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("WORDWRAP", (0, 0), (-1, -1), True),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al exportar bitácora a PDF: {str(e)}")
    finally:
        if conexion and conexion.is_connected() and cursor is not None:
            cursor.close()
            conexion.close()
