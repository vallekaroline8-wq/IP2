from io import BytesIO
from pathlib import Path
from datetime import datetime

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from mysql.connector import Error

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
    XDRPositiveSize2D
)
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins

from database.conexion import get_connection


def exportar_ips_excel():
    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                ip.direccion_ip AS direccion,
                seg.nombre AS segmento,
                est.nombre AS estado,
                (
                    SELECT eq.nombre_equipo
                    FROM tbl_asignacion_ip asig
                    LEFT JOIN tbl_equipo eq
                        ON eq.id_equipo = asig.id_equipo
                    WHERE asig.id_ip = ip.id_ip
                      AND asig.fecha_liberacion IS NULL
                    ORDER BY asig.fecha_asignacion DESC
                    LIMIT 1
                ) AS equipo
            FROM tbl_ip ip
            LEFT JOIN tbl_segmento seg
                ON seg.id_segmento = ip.id_segmento
            LEFT JOIN tbl_estado est
                ON est.id_estado = ip.id_estado
            WHERE ip.id_estado IN (3, 4, 5)
            ORDER BY ip.direccion_ip ASC
        """)

        datos = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Direcciones IP"
        ws.sheet_view.showGridLines = True

        azul = "0A4B8F"
        blanco = "FFFFFF"
        fill_azul = PatternFill("solid", fgColor=azul)
        fill_blanco = PatternFill("solid", fgColor=blanco)

        # Configuración de anchos según requerimientos
        max_col_banner = 5  # A-E para mantener simetría con el encabezado
        max_col_tabla = 4   # A-D únicamente para los 4 campos reales

        # --- ALINEACIÓN Y ESTRUCTURA DEL ENCABEZADO Y LOGO ---
        ws.column_dimensions["A"].width = 22

        # Alturas de filas para el banner del encabezado
        ws.row_dimensions[1].height = 12
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 24
        ws.row_dimensions[5].height = 12

        # Combinar espacio para el logo (A1:A4)
        ws.merge_cells("A1:A4")
        for row in ws["A1:A4"]:
            for cell in row:
                cell.fill = fill_blanco

        logo_path = Path(__file__).parent.parent / "assets" / "hospital_logo.png"

        if logo_path.exists():
            imagen = Image(str(logo_path))
            imagen.width = 110
            imagen.height = 110

            marker = AnchorMarker(
                col=0,
                row=0,
                colOff=pixels_to_EMU(18),
                rowOff=pixels_to_EMU(18)
            )

            imagen.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(
                    pixels_to_EMU(imagen.width),
                    pixels_to_EMU(imagen.height)
                )
            )
            ws.add_image(imagen)

        # ---------------- ENCABEZADO DEL REPORTE ----------------
        # Banner azul extending hasta max_col_banner (E)
        for fila in range(2, 5):
            for columna in range(2, max_col_banner + 1):
                ws.cell(row=fila, column=columna).fill = fill_azul

        # Título principal
        ws.merge_cells(f"B2:{get_column_letter(max_col_banner)}2")
        ws["B2"] = "HOSPITAL MILITAR"
        ws["B2"].font = Font(bold=True, size=16, color=blanco)
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        # Subtítulo
        ws.merge_cells(f"B3:{get_column_letter(max_col_banner)}3")
        ws["B3"] = "SISTEMA SIGIP"
        ws["B3"].font = Font(bold=True, size=12, color=blanco)
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

        # Nombre del reporte
        ws.merge_cells(f"B4:{get_column_letter(max_col_banner)}4")
        ws["B4"] = "REPORTE GENERAL DE DIRECCIONES IP"
        ws["B4"].font = Font(bold=True, size=11, color=blanco)
        ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

        # --- METADATOS DEL REPORTE ---
        fecha = datetime.now()
        
        ws["A6"] = "Fecha:"
        ws["B6"] = fecha.strftime("%d/%m/%Y")
        ws["A7"] = "Hora:"
        ws["B7"] = fecha.strftime("%H:%M")

        ws["C6"] = "Total Direcciones IP:"
        ws["D6"] = len(datos)
        ws["C7"] = "Registros:"
        ws["D7"] = len(datos)

        for celda in ["A6", "A7", "C6", "C7"]:
            ws[celda].font = Font(bold=True)
            ws[celda].alignment = Alignment(horizontal="left")

        ws["B6"].alignment = Alignment(horizontal="left")
        ws["B7"].alignment = Alignment(horizontal="left")
        ws["D6"].alignment = Alignment(horizontal="right")
        ws["D7"].alignment = Alignment(horizontal="right")

        # --- TABLA DE DATOS ---
        encabezados = ["Dirección IP", "Segmento", "Estado", "Equipo"]
        fila_inicio = 10

        borde = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # Encabezados de la tabla
        for columna, texto in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_inicio, column=columna)
            celda.value = texto
            celda.font = Font(bold=True, color=blanco)
            celda.fill = fill_azul
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = borde

        # Llenar datos
        fila = fila_inicio + 1
        for item in datos:
            ws.cell(fila, 1).value = item["direccion"]
            ws.cell(fila, 2).value = item["segmento"]
            ws.cell(fila, 3).value = item["estado"]
            ws.cell(fila, 4).value = item["equipo"] or "-"

            for columna in range(1, max_col_tabla + 1):
                celda = ws.cell(fila, columna)
                celda.border = borde
                celda.alignment = Alignment(
                    vertical="center", 
                    horizontal="center" if columna in (1, 3) else "left"
                )

            fila += 1

        ultima_fila = max(fila - 1, fila_inicio)

        # Crear tabla oficial de Excel acotada estrictamente a max_col_tabla (A-D)
        if datos:
            tabla = Table(
                displayName="TablaDireccionesIP",
                ref=f"A{fila_inicio}:{get_column_letter(max_col_tabla)}{ultima_fila}"
            )
            estilo = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)

        # Congelar paneles debajo de los encabezados
        ws.freeze_panes = f"A{fila_inicio + 1}"

        # Ajuste dinámico de ancho de columnas (limitado a max_col_tabla)
        for columna in range(1, max_col_tabla + 1):
            letra = get_column_letter(columna)
            longitud = 0
            for fila_excel in range(fila_inicio, ultima_fila + 1):
                valor = ws.cell(fila_excel, columna).value
                if valor is not None:
                    longitud = max(longitud, len(str(valor)))
            ws.column_dimensions[letra].width = max(longitud + 5, 18)

        # Pie de página / Resumen (centrado cubriendo el ancho del banner)
        fila_resumen = ultima_fila + 2
        ws.merge_cells(
            start_row=fila_resumen,
            start_column=1,
            end_row=fila_resumen,
            end_column=max_col_banner
        )

        celda_resumen = ws.cell(row=fila_resumen, column=1)
        celda_resumen.value = "Reporte generado automáticamente por SIGIP"
        celda_resumen.font = Font(italic=True, size=10, color="666666")
        celda_resumen.alignment = Alignment(horizontal="center")

        # Configuración de impresión
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_margins = PageMargins(
            left=0.5,
            right=0.5,
            top=0.5,
            bottom=0.5
        )

        wb.properties.creator = "SIGIP"
        wb.properties.title = "Reporte General de Direcciones IP"
        wb.properties.subject = "Hospital Militar"
        wb.properties.company = "Hospital Militar"
        wb.properties.category = "Reportes"

        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)

        return StreamingResponse(
            archivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=Reporte_Direcciones_IP_SIGIP.xlsx"
            }
        )

    except Error as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error en la base de datos al exportar direcciones IP: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error inesperado al generar el archivo Excel: {str(e)}"
        )

    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        if conexion and conexion.is_connected():
            try:
                conexion.close()
            except Exception:
                pass