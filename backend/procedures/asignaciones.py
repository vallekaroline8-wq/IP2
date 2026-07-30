from io import BytesIO
from pathlib import Path
from datetime import datetime
import os

from database.conexion import get_connection
from procedures.bitacoramodulo import registrar_bitacora

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from mysql.connector import Error

# ==========================
# OPENPYXL
# ==========================
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
    XDRPositiveSize2D
)
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from openpyxl.worksheet.page import PageMargins

# ==========================
# REPORTLAB
# ==========================
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate,
    Table as PDFTable,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage
)


# ==========================================
# CANVAS PERSONALIZADO PARA NUMERACIÓN PDF
# ==========================================

class NumberedCanvas(canvas.Canvas):
    """
    Canvas personalizado para calcular el total de páginas y dibujar
    el pie de página (con numeración dinámicamente) en dos pasadas.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#555555"))

        # Pie de página (Izquierda: Nombre del Hospital, Derecha: Páginas)
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        texto_izq = f"Hospital Militar — Generado: {fecha_actual}"
        texto_der = f"Página {self._pageNumber} de {page_count}"

        self.drawString(1.5 * cm, 0.8 * cm, texto_izq)
        self.drawRightString(26.4 * cm, 0.8 * cm, texto_der)

        # Línea divisoria inferior
        self.setStrokeColor(colors.HexColor("#CCCCCC"))
        self.setLineWidth(0.5)
        self.line(1.5 * cm, 1.2 * cm, 26.4 * cm, 1.2 * cm)

        self.restoreState()


# ==========================================
# LISTAR ASIGNACIONES
# ==========================================

def obtener_asignaciones(page: int = 1):
    """
    Obtiene únicamente las direcciones IP que se encuentran asignadas.
    """

    if page < 1:
        page = 1

    conexion = get_connection()

    try:

        cursor = conexion.cursor(dictionary=True)

        page_size = 10
        offset = (page - 1) * page_size

        consulta_sql = """
            SELECT
                ai.id_asignacion AS id,
                ip.direccion_ip AS ip_direccion,
                e.nombre_equipo AS equipo_nombre,
                ai.fecha_asignacion,
                ai.fecha_liberacion,
                ai.id_estado,
                est.nombre AS estado,

                CASE
                    WHEN ai.id_estado = 4 THEN TRUE
                    ELSE FALSE
                END AS activo

            FROM tbl_asignacion_ip ai

            INNER JOIN tbl_ip ip
                ON ai.id_ip = ip.id_ip

            INNER JOIN tbl_equipo e
                ON ai.id_equipo = e.id_equipo

            INNER JOIN tbl_estado est
                ON ai.id_estado = est.id_estado

            WHERE ai.id_estado = 4

            ORDER BY ai.id_asignacion DESC

            LIMIT %s, %s
        """

        cursor.execute(
            consulta_sql,
            (
                offset,
                page_size
            )
        )

        items = cursor.fetchall()

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM tbl_asignacion_ip
            WHERE id_estado = 4
        """)

        total = cursor.fetchone()["total"]

        pages = (total + page_size - 1) // page_size

        return {
            "items": items,
            "pages": pages,
            "total": total
        }

    except Error as e:

        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener asignaciones: {str(e)}"
        )

    finally:

        if conexion.is_connected():
            cursor.close()
            conexion.close()

# ==========================================
# LIBERAR ASIGNACIÓN DE IP
# ==========================================

def liberar_asignacion(id_asignacion: int, id_usuario_actual: int):
    """
    Libera una dirección IP asignada.
    """

    conexion = get_connection()

    try:

        cursor = conexion.cursor(dictionary=True)

        # Buscar asignación
        cursor.execute("""
            SELECT
                ai.id_ip,
                ai.id_estado,
                ip.direccion_ip,
                eq.nombre_equipo
            FROM tbl_asignacion_ip ai
            INNER JOIN tbl_ip ip
            ON ip.id_ip = ai.id_ip
            INNER JOIN tbl_equipo eq
            ON eq.id_equipo = ai.id_equipo
            WHERE ai.id_asignacion = %s
        """, (id_asignacion,))

        asignacion = cursor.fetchone()

        if not asignacion:
            raise HTTPException(
                status_code=404,
                detail="La asignación no existe."
            )

        # Verificar si ya está liberada
        if asignacion["id_estado"] == 3:
            raise HTTPException(
                status_code=400,
                detail="La asignación ya fue liberada."
            )

        # Actualizar asignación
        cursor.execute("""
            UPDATE tbl_asignacion_ip
            SET
                id_estado = 3,
                fecha_liberacion = NOW()
            WHERE id_asignacion = %s
        """, (id_asignacion,))

        # Actualizar IP
        cursor.execute("""
            UPDATE tbl_ip
            SET id_estado = 3
            WHERE id_ip = %s
        """, (asignacion["id_ip"],))

        conexion.commit()

        registrar_bitacora(
            id_usuario=id_usuario_actual,
            accion="LIBERAR",
            tabla_afectada="tbl_asignacion_ip",
            registro_id=id_asignacion,
            detalle=(
            f"Se liberó la dirección IP "
            f"'{asignacion['direccion_ip']}' "
            f"del equipo "
            f"'{asignacion['nombre_equipo']}'."
            )
        )

        return {
            "mensaje": "Dirección IP liberada correctamente."
        }

    except Error as e:

        conexion.rollback()

        raise HTTPException(
            status_code=500,
            detail=f"Error al liberar la dirección IP: {str(e)}"
        )

    finally:

        if conexion.is_connected():
            cursor.close()
            conexion.close()

# ==========================================
# COMBO EQUIPOS
# ==========================================

def obtener_equipos():

    conexion = get_connection()

    try:
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                id_equipo AS id,
                nombre_equipo AS nombre
            FROM tbl_equipo
            WHERE id_estado = 1
            ORDER BY nombre_equipo
        """)

        return cursor.fetchall()

    except Error as e:
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )

    finally:
        cursor.close()
        conexion.close()


# ==========================================
# COMBO SEGMENTOS
# ==========================================

def obtener_segmentos():

    conexion = get_connection()

    try:
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                id_segmento AS id,
                nombre
            FROM tbl_segmento
            WHERE id_estado = 1
            ORDER BY nombre
        """)

        return cursor.fetchall()

    except Error as e:
        raise HTTPException(
            status_code=500,
            detail=str(e)
        )

    finally:
        cursor.close()
        conexion.close()

# ==========================================
# IPS DISPONIBLES
# ==========================================

def obtener_ips_disponibles(id_segmento):
    """
    Obtiene las IP disponibles de un segmento.
    """

    conexion = get_connection()

    try:

        cursor = conexion.cursor(dictionary=True)

        print("ID SEGMENTO RECIBIDO:", id_segmento)

        cursor.execute("""
            SELECT
                id_ip AS id,
                direccion_ip AS direccion
            FROM tbl_ip
            WHERE id_segmento = %s
              AND id_estado = 3
            ORDER BY direccion_ip
        """, (id_segmento,))

        return cursor.fetchall()

    except Error as e:

        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener las IP disponibles: {str(e)}"
        )

    finally:

        if conexion.is_connected():
            cursor.close()
            conexion.close()

# ==========================================
# ASIGNAR DIRECCIÓN IP
# ==========================================

def asignar_ip(id_ip: int, id_equipo: int, id_usuario: int, id_usuario_actual: int):
    """
    Asigna una dirección IP a un equipo.
    """

    conexion = get_connection()

    try:

        cursor = conexion.cursor(dictionary=True)

        # ======================================
        # Verificar que la IP exista
        # ======================================

        cursor.execute("""
            SELECT
                id_ip,
                direccion_ip,
                id_estado
            FROM tbl_ip
            WHERE id_ip = %s
        """, (id_ip,))

        ip = cursor.fetchone()

        if not ip:
            raise HTTPException(
                status_code=404,
                detail="La dirección IP no existe."
            )

        # Debe estar DISPONIBLE
        if ip["id_estado"] != 3:
            raise HTTPException(
                status_code=400,
                detail="La dirección IP no está disponible."
            )

        # ======================================
        # Verificar equipo
        # ======================================

        cursor.execute("""
            SELECT id_equipo, nombre_equipo
            FROM tbl_equipo
            WHERE id_equipo = %s
        """, (id_equipo,))

        equipo = cursor.fetchone()

        if not equipo:
            raise HTTPException(
                status_code=404,
                detail="El equipo no existe."
            )

        # ======================================
        # Verificar usuario
        # ======================================

        cursor.execute("""
            SELECT id_usuario
            FROM tbl_usuario
            WHERE id_usuario = %s
        """, (id_usuario,))

        if cursor.fetchone() is None:
            raise HTTPException(
                status_code=404,
                detail="El usuario no existe."
            )

        # ======================================
        # Registrar asignación
        # ======================================

        cursor.execute("""
            INSERT INTO tbl_asignacion_ip
            (
                id_ip,
                id_equipo,
                id_usuario,
                fecha_asignacion,
                id_estado
            )
            VALUES
            (
                %s,
                %s,
                %s,
                NOW(),
                %s
            )
        """, (
            id_ip,
            id_equipo,
            id_usuario,
            4
        ))

        # ======================================
        # Cambiar estado de la IP
        # ======================================

        cursor.execute("""
            UPDATE tbl_ip
            SET id_estado = 4
            WHERE id_ip = %s
        """, (id_ip,))

        conexion.commit()

        registrar_bitacora(
            id_usuario=id_usuario_actual,
            accion="ASIGNAR",
            tabla_afectada="tbl_asignacion_ip",
            registro_id=cursor.lastrowid,
            detalle=(
            f"Se asignó la dirección IP "
            f"'{ip['direccion_ip']}' "
            f"al equipo "
            f"'{equipo['nombre_equipo']}'."
            )
        )

        return {
            "mensaje": "Dirección IP asignada correctamente."
        }

    except Error as e:

        conexion.rollback()

        raise HTTPException(
            status_code=500,
            detail=f"Error al asignar la dirección IP: {str(e)}"
        )

    finally:

        if conexion.is_connected():
            cursor.close()
            conexion.close()
    """
    Asigna una dirección IP a un equipo.
    """

    conexion = get_connection()

    try:

        cursor = conexion.cursor(dictionary=True)

        # ======================================
        # Verificar que la IP exista
        # ======================================

        cursor.execute("""
            SELECT
                id_ip,
                id_estado
            FROM tbl_ip
            WHERE id_ip = %s
        """, (id_ip,))

        ip = cursor.fetchone()

        if not ip:
            raise HTTPException(
                status_code=404,
                detail="La dirección IP no existe."
            )

        # id_estado = 3 -> DISPONIBLE
        if ip["id_estado"] != 3:
            raise HTTPException(
                status_code=400,
                detail="La dirección IP no está disponible."
            )

        # ======================================
        # Verificar que exista el equipo
        # ======================================

        cursor.execute("""
            SELECT id_equipo
            FROM tbl_equipo
            WHERE id_equipo = %s
        """, (id_equipo,))

        if cursor.fetchone() is None:
            raise HTTPException(
                status_code=404,
                detail="El equipo no existe."
            )

        # ======================================
        # Verificar que exista el usuario
        # ======================================

        cursor.execute("""
            SELECT id_usuario
            FROM tbl_usuario
            WHERE id_usuario = %s
        """, (id_usuario,))

        if cursor.fetchone() is None:
            raise HTTPException(
                status_code=404,
                detail="El usuario no existe."
            )

        # ======================================
        # Registrar asignación
        # ======================================

        cursor.execute("""
            INSERT INTO tbl_asignacion_ip
            (
                id_ip,
                id_equipo,
                id_usuario,
                fecha_asignacion,
                estado_asignacion
            )
            VALUES
            (
                %s,
                %s,
                %s,
                NOW(),
                'ACTIVA'
            )
        """, (
            id_ip,
            id_equipo,
            id_usuario
        ))

        # ======================================
        # Cambiar estado de la IP
        # id_estado = 4 -> ASIGNADA
        # ======================================

        cursor.execute("""
            UPDATE tbl_ip
            SET id_estado = 4
            WHERE id_ip = %s
        """, (id_ip,))

        conexion.commit()

        return {
            "mensaje": "Dirección IP asignada correctamente."
        }

    except Error as e:

        conexion.rollback()

        raise HTTPException(
            status_code=500,
            detail=f"Error al asignar la dirección IP: {str(e)}"
        )

    finally:

        if conexion.is_connected():
            cursor.close()
            conexion.close()

# ======================================
# IMPRIMIR EXEL
# ======================================

def exportar_asignaciones_excel():
    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                ip.direccion_ip,
                e.nombre_equipo,
                ai.fecha_asignacion,
                CASE
                    WHEN ai.id_estado = 4 THEN 'Activa'
                    WHEN ai.id_estado = 3 THEN 'Liberada'
                    ELSE 'Desconocida'
                END AS estado
            FROM tbl_asignacion_ip ai
            INNER JOIN tbl_ip ip
                ON ip.id_ip = ai.id_ip
            INNER JOIN tbl_equipo e
                ON e.id_equipo = ai.id_equipo
            ORDER BY ai.fecha_asignacion DESC
        """)

        datos = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Asignaciones"
        ws.sheet_view.showGridLines = True

        azul = "0A4B8F"
        blanco = "FFFFFF"

        # ---------------- ENCABEZADO ----------------

        ws.column_dimensions["A"].width = 20

        ws.row_dimensions[1].height = 12
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 24
        ws.row_dimensions[5].height = 12

        ws.merge_cells("A1:A4")
        fill_blanco = PatternFill("solid", fgColor=blanco)

        for row in ws["A1:A4"]:
            for cell in row:
                cell.fill = fill_blanco

        logo = Path(__file__).parent.parent / "assets" / "hospital_logo.png"

        if logo.exists():
            imagen = Image(str(logo))
            imagen.width = 100
            imagen.height = 100

            marker = AnchorMarker(
                col=0,
                row=0,
                colOff=pixels_to_EMU(16),
                rowOff=pixels_to_EMU(18)
            )

            imagen.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(
                    pixels_to_EMU(imagen.width),
                    pixels_to_EMU(imagen.height)
                )
            )

            ws.add_image(imagen)

        fill_azul = PatternFill("solid", fgColor=azul)

        for fila in range(2, 5):
            for columna in range(2, 6):
                ws.cell(row=fila, column=columna).fill = fill_azul

        ws.merge_cells("B2:E2")
        ws["B2"] = "HOSPITAL MILITAR"
        ws["B2"].font = Font(bold=True, size=16, color=blanco)
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B3:E3")
        ws["B3"] = "SISTEMA SIGIP"
        ws["B3"].font = Font(bold=True, size=12, color=blanco)
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("B4:E4")
        ws["B4"] = "REPORTE DE ASIGNACIONES DE IP"
        ws["B4"].font = Font(bold=True, size=11, color=blanco)
        ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

        # ---------------- METADATOS ----------------

        fecha = datetime.now()

        ws["B6"] = "Fecha"
        ws["C6"] = fecha.strftime("%d/%m/%Y")

        ws["B7"] = "Hora"
        ws["C7"] = fecha.strftime("%H:%M")

        ws["D6"] = "Total Registros"
        ws["E6"] = len(datos)

        ws["D7"] = "Asignaciones"
        ws["E7"] = len(datos)

        for celda in ["B6", "B7", "D6", "D7"]:
            ws[celda].font = Font(bold=True)

        # ---------------- TABLA ----------------

        encabezados = [
            "Dirección IP",
            "Equipo",
            "Fecha Asignación",
            "Estado"
        ]

        fila_inicio = 10

        borde = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        for columna, texto in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_inicio, column=columna)
            celda.value = texto
            celda.font = Font(bold=True, color=blanco)
            celda.fill = PatternFill(fill_type="solid", fgColor=azul)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = borde

        fila = fila_inicio + 1

        for asignacion in datos:

            fecha_asignacion = ""

            if asignacion["fecha_asignacion"]:
                fecha_asignacion = asignacion["fecha_asignacion"].strftime("%d/%m/%Y %H:%M")

            ws.cell(fila, 1).value = asignacion["direccion_ip"]
            ws.cell(fila, 2).value = asignacion["nombre_equipo"]
            ws.cell(fila, 3).value = fecha_asignacion
            ws.cell(fila, 4).value = asignacion["estado"]

            for columna in range(1, 5):
                celda = ws.cell(fila, columna)
                celda.border = borde
                celda.alignment = Alignment(vertical="center")

            fila += 1

        ultima_fila = max(fila - 1, fila_inicio)

        if datos:
            tabla = ExcelTable(
                displayName="TablaAsignaciones",
                ref=f"A{fila_inicio}:D{ultima_fila}"
            )

            estilo = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False
            )

            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)

        ws.freeze_panes = f"A{fila_inicio + 1}"

        for columna in range(1, 5):
            letra = get_column_letter(columna)
            longitud = 0

            for fila_excel in range(1, ultima_fila + 1):
                valor = ws.cell(fila_excel, columna).value

                if valor is not None:
                    longitud = max(longitud, len(str(valor)))

            ws.column_dimensions[letra].width = max(longitud + 4, 18)

        fila_resumen = ultima_fila + 3

        ws.merge_cells(
            start_row=fila_resumen,
            start_column=1,
            end_row=fila_resumen,
            end_column=4
        )

        celda_resumen = ws.cell(row=fila_resumen, column=1)
        celda_resumen.value = "Reporte generado automáticamente por SIGIP"
        celda_resumen.font = Font(
            italic=True,
            size=10,
            color="666666"
        )
        celda_resumen.alignment = Alignment(horizontal="center")

        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1

        ws.page_margins = PageMargins(
            left=0.3,
            right=0.3,
            top=0.5,
            bottom=0.5
        )

        wb.properties.creator = "SIGIP"
        wb.properties.title = "Reporte de Asignaciones"
        wb.properties.subject = "Hospital Militar"

        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)

        return StreamingResponse(
            archivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=Reporte_Asignaciones_SIGIP.xlsx"
            }
        )

    except Error as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al exportar asignaciones: {str(e)}"
        )

    finally:
        if cursor:
            cursor.close()

        if conexion and conexion.is_connected():
            conexion.close()

# ======================================
# IMPRIMIR PDF
# ======================================
import os
from datetime import datetime

from mysql.connector import Error

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate,
    Table as PDFTable,
    TableStyle,
    Paragraph,
    Spacer,
    Image as RLImage
)
from database.conexion import get_connection


class NumberedCanvas(canvas.Canvas):
    """
    Canvas personalizado para numeración de páginas.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)

        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()

        super().save()

    def draw_page_decorations(self, page_count):

        self.saveState()

        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#555555"))

        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")

        self.drawString(
            1.5 * cm,
            0.8 * cm,
            f"Hospital Militar — Generado: {fecha_actual}"
        )

        self.drawRightString(
            26.4 * cm,
            0.8 * cm,
            f"Página {self._pageNumber} de {page_count}"
        )

        self.setStrokeColor(colors.HexColor("#CCCCCC"))
        self.setLineWidth(0.5)

        self.line(
            1.5 * cm,
            1.2 * cm,
            26.4 * cm,
            1.2 * cm
        )

        self.restoreState()


def exportar_asignaciones_pdf():

    conexion = None
    cursor = None

    try:

        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT

                ai.id_asignacion AS id,

                ip.direccion_ip AS ip_direccion,

                e.nombre_equipo AS equipo_nombre,

                ai.fecha_asignacion,

                ai.fecha_liberacion,

                ai.id_estado,

                est.nombre AS estado,

                CASE
                    WHEN ai.id_estado = 4 THEN TRUE
                    ELSE FALSE
                END AS activo

            FROM tbl_asignacion_ip ai

            INNER JOIN tbl_ip ip
                ON ai.id_ip = ip.id_ip

            INNER JOIN tbl_equipo e
                ON ai.id_equipo = e.id_equipo

            INNER JOIN tbl_estado est
                ON ai.id_estado = est.id_estado

            WHERE ai.id_estado = 4

            ORDER BY ai.id_asignacion DESC
        """)

        asignaciones = cursor.fetchall()

        carpeta = "exports"

        if not os.path.exists(carpeta):
            os.makedirs(carpeta)

        archivo_pdf = os.path.join(
            carpeta,
            "Reporte_Asignaciones_IP.pdf"
        )

        documento = SimpleDocTemplate(
            archivo_pdf,
            pagesize=landscape(letter),
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.8 * cm
        )

        estilos = getSampleStyleSheet()

        titulo_estilo = ParagraphStyle(
            "TituloSIGIP",
            parent=estilos["Title"],
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#003366")
        )

        estilo_celda = ParagraphStyle(
            "Celda",
            parent=estilos["Normal"],
            fontSize=8,
            leading=10,
            alignment=TA_LEFT
        )

        estilo_encabezado = ParagraphStyle(
            "Encabezado",
            parent=estilos["Normal"],
            fontSize=9,
            leading=11,
            textColor=colors.white,
            fontName="Helvetica-Bold",
            alignment=TA_CENTER
        )

        contenido = []

        logo_path = os.path.join(
            "assets",
            "hospital_logo.png"
        )

        titulo = """
        <b>HOSPITAL MILITAR</b><br/>
        <font size=10 color="#555555">
        SIGIP - Sistema de Gestión de Direcciones IP
        </font><br/>
        <font size=12 color="#003366">
        <b>REPORTE DE ASIGNACIONES DE IP</b>
        </font>
        """

        paragraph = Paragraph(
            titulo,
            titulo_estilo
        )

        if os.path.exists(logo_path):

            logo = RLImage(
                logo_path,
                width=2.3 * cm,
                height=2.3 * cm
            )

            header = PDFTable(
                [[logo, paragraph]],
                colWidths=[2.8 * cm, 22.1 * cm]
            )

            header.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (0, 0), "LEFT"),
                ("ALIGN", (1, 0), (1, 0), "CENTER"),
            ]))

            contenido.append(header)

        else:

            contenido.append(paragraph)

        contenido.append(
            Spacer(1, 0.6 * cm)
        )

        datos = [[
            Paragraph("Dirección IP", estilo_encabezado),
            Paragraph("Equipo", estilo_encabezado),
            Paragraph("Fecha Asignación", estilo_encabezado),
            Paragraph("Estado", estilo_encabezado),
        ]]

        for fila in asignaciones:


            fecha = ""

            if fila["fecha_asignacion"]:
                fecha = fila["fecha_asignacion"].strftime("%d/%m/%Y %H:%M")

            datos.append([
                Paragraph(str(fila["ip_direccion"]), estilo_celda),
                Paragraph(str(fila["equipo_nombre"]), estilo_celda),
                Paragraph(fecha, estilo_celda),
                Paragraph(str(fila["estado"]), estilo_celda),
            ])

            tabla = PDFTable(
                datos,
                repeatRows=1,
                colWidths=[
                5 * cm,
                10 * cm,
                6 * cm,
                4 * cm
                ]
            )

        tabla.setStyle(TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F9F9F9")),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6)

        ]))

        contenido.append(tabla)

        documento.build(
            contenido,
            canvasmaker=NumberedCanvas
        )

        return archivo_pdf

    except Error as e:
        raise Exception(
            f"Error al consultar la base de datos: {str(e)}"
        )

    except Exception as e:
        raise Exception(
            f"Error al generar el PDF: {str(e)}"
        )

    finally:

        if cursor:
            cursor.close()

        if conexion and conexion.is_connected():
            conexion.close()