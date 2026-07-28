from io import BytesIO
from pathlib import Path
from datetime import datetime

from fastapi import HTTPException
from fastapi.responses import StreamingResponse

from mysql.connector import Error

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker,
    OneCellAnchor,
    XDRPositiveSize2D
)
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins

from database.conexion import get_connection


def exportar_equipos_excel():
    conexion = None
    cursor = None

    try:
        conexion = get_connection()
        cursor = conexion.cursor(dictionary=True)

        cursor.execute("""
            SELECT
                e.nombre_equipo,
                td.nombre AS tipo,
                e.marca,
                e.modelo,
                d.nombre AS departamento,
                e.ubicacion,
                e.area,
                e.extension
            FROM tbl_equipo e
            INNER JOIN tbl_tipo_dispositivo td
                ON td.id_tipo = e.id_tipo
            INNER JOIN tbl_departamento d
                ON d.id_departamento = e.id_departamento
            WHERE e.id_estado <> 6
            ORDER BY e.nombre_equipo ASC
        """)

        datos = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Equipos"
        ws.sheet_view.showGridLines = True

        azul = "0A4B8F"
        blanco = "FFFFFF"

        # --- ALINEACIÓN Y ESTRUCTURA DEL ENCABEZADO Y LOGO ---
        ws.column_dimensions["A"].width = 20

        # Ajuste de alturas para que el banner azul se vea simétrico con el logo
        ws.row_dimensions[1].height = 12
        ws.row_dimensions[2].height = 28
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 24
        ws.row_dimensions[5].height = 12

        # Combinar el espacio del logo A1:A4 y forzar fondo blanco
        ws.merge_cells("A1:A4")
        fill_blanco = PatternFill("solid", fgColor=blanco)
        for row in ws["A1:A4"]:
            for cell in row:
                cell.fill = fill_blanco

        logo = Path(__file__).parent.parent / "assets" / "hospital_logo.png"

        if logo.exists():
            imagen = Image(str(logo))

            # Dimensiones del logo
            imagen.width = 180
            imagen.height = 180

            # Posicionamiento preciso dentro de la celda combinada
            marker = AnchorMarker(
                col=0,                      # Columna A
                row=0,                      # Fila 1
                colOff=pixels_to_EMU(35),   # Desplazamiento horizontal
                rowOff=pixels_to_EMU(8)     # Desplazamiento vertical
            )

            imagen.anchor = OneCellAnchor(
                _from=marker,
                ext=XDRPositiveSize2D(
                    pixels_to_EMU(imagen.width),
                    pixels_to_EMU(imagen.height)
                )
            )

            ws.add_image(imagen)

        # ---------------- ENCABEZADO DEL REPORTE ----------------
        fill_azul = PatternFill("solid", fgColor=azul)

        # Pintar el banner azul de la columna B a la H (Columnas 2 a 8)
        for fila in range(2, 5):
            for columna in range(2, 9):  # B -> H
                ws.cell(row=fila, column=columna).fill = fill_azul

        # Título principal
        ws.merge_cells("B2:H2")
        ws["B2"] = "HOSPITAL MILITAR"
        ws["B2"].font = Font(bold=True, size=16, color=blanco)
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

        # Subtítulo
        ws.merge_cells("B3:H3")
        ws["B3"] = "SISTEMA SIGIP"
        ws["B3"].font = Font(bold=True, size=12, color=blanco)
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

        # Nombre del reporte
        ws.merge_cells("B4:H4")
        ws["B4"] = "REPORTE GENERAL DE EQUIPOS"
        ws["B4"].font = Font(bold=True, size=11, color=blanco)
        ws["B4"].alignment = Alignment(horizontal="center", vertical="center")

        # --- METADATOS DEL REPORTE ---
        fecha = datetime.now()
        ws["B6"] = "Fecha"
        ws["C6"] = fecha.strftime("%d/%m/%Y")

        ws["B7"] = "Hora"
        ws["C7"] = fecha.strftime("%H:%M")

        ws["E6"] = "Total Equipos"
        ws["F6"] = len(datos)

        ws["E7"] = "Registros"
        ws["F7"] = len(datos)

        for celda in ["B6", "B7", "E6", "E7"]:
            ws[celda].font = Font(bold=True)

        # Configuración de alineación de Metadatos
        ws["B6"].alignment = Alignment(horizontal="left")
        ws["B7"].alignment = Alignment(horizontal="left")
        ws["C6"].alignment = Alignment(horizontal="left")
        ws["C7"].alignment = Alignment(horizontal="left")
        ws["E6"].alignment = Alignment(horizontal="left")
        ws["E7"].alignment = Alignment(horizontal="left")
        ws["F6"].alignment = Alignment(horizontal="right")
        ws["F7"].alignment = Alignment(horizontal="right")

        # --- TABLA DE DATOS ---
        encabezados = [
            "Nombre",
            "Tipo",
            "Marca",
            "Modelo",
            "Departamento",
            "Ubicación",
            "Área",
            "Extensión"
        ]

        fila_inicio = 10

        borde = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # Insertar encabezados
        for columna, texto in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila_inicio, column=columna)
            celda.value = texto
            celda.font = Font(bold=True, color=blanco)
            celda.fill = PatternFill(fill_type="solid", fgColor=azul)
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.border = borde

        # Insertar filas
        fila = fila_inicio + 1
        for equipo in datos:
            ws.cell(fila, 1).value = equipo["nombre_equipo"]
            ws.cell(fila, 2).value = equipo["tipo"]
            ws.cell(fila, 3).value = equipo["marca"]
            ws.cell(fila, 4).value = equipo["modelo"]
            ws.cell(fila, 5).value = equipo["departamento"]
            ws.cell(fila, 6).value = equipo["ubicacion"]
            ws.cell(fila, 7).value = equipo["area"]
            ws.cell(fila, 8).value = equipo["extension"]

            for columna in range(1, 9):
                celda = ws.cell(fila, columna)
                celda.border = borde
                celda.alignment = Alignment(vertical="center")

            fila += 1

        ultima_fila = max(fila - 1, fila_inicio)

        # Crear tabla oficial de Excel
        if datos:
            tabla = Table(
                displayName="TablaEquipos",
                ref=f"A{fila_inicio}:H{ultima_fila}"
            )
            estilo = TableStyleInfo(
                name="TableStyleLight1",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)

        ws.freeze_panes = f"A{fila_inicio + 1}"

        # Ajuste de ancho de columnas
        for columna in range(1, 9):
            letra = get_column_letter(columna)
            longitud = 0
            for fila_excel in range(1, ultima_fila + 1):
                valor = ws.cell(fila_excel, columna).value
                if valor is not None:
                    longitud = max(longitud, len(str(valor)))
            # Mantener un ancho mínimo cómodo
            ws.column_dimensions[letra].width = max(longitud + 4, 15)

        # Pie de página / Resumen
        fila_resumen = ultima_fila + 3
        ws.merge_cells(
            start_row=fila_resumen,
            start_column=1,
            end_row=fila_resumen,
            end_column=8
        )

        celda_resumen = ws.cell(row=fila_resumen, column=1)
        celda_resumen.value = "Reporte generado automáticamente por SIGIP"
        celda_resumen.font = Font(italic=True, size=10, color="666666")
        celda_resumen.alignment = Alignment(horizontal="center")

        # Configuración de impresión
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_margins = PageMargins(
            left=0.3,
            right=0.3,
            top=0.5,
            bottom=0.5
        )

        wb.properties.creator = "SIGIP"
        wb.properties.title = "Reporte General de Equipos"
        wb.properties.subject = "Hospital Militar"
        wb.properties.company = "Hospital Militar"
        wb.properties.category = "Reportes"

        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)

        return StreamingResponse(
            archivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=Reporte_Equipos_SIGIP.xlsx"
            }
        )

    except Error as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al exportar equipos: {str(e)}"
        )

    finally:
        if cursor:
            cursor.close()
        if conexion and conexion.is_connected():
            conexion.close()