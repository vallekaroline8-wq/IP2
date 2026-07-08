"""Generación de reportes PDF y Excel para SIGIP."""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def build_excel(title: str, columns: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
    tcell = ws.cell(row=1, column=1, value=f"HOSPITAL MILITAR - SIGIP | {title}")
    tcell.font = Font(bold=True, size=14, color="FFFFFF")
    tcell.fill = PatternFill("solid", fgColor="0A4B8F")
    tcell.alignment = Alignment(horizontal="center")
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=c, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="083D75")
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=3):
        for c, key in enumerate(columns, start=1):
            ws.cell(row=r, column=c, value=str(row.get(key, "")))
    for c in range(1, len(columns) + 1):
        ws.column_dimensions[chr(64 + c)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(title: str, columns: list, rows: list) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle("h", parent=styles["Title"], textColor=colors.HexColor("#0A4B8F"), fontSize=16)
    sub = ParagraphStyle("s", parent=styles["Normal"], textColor=colors.HexColor("#64748B"), fontSize=9)
    elements = [
        Paragraph("HOSPITAL MILITAR &mdash; SIGIP", header_style),
        Paragraph(title, styles["Heading2"]),
        Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub),
        Spacer(1, 8),
    ]
    data = [columns] + [[str(row.get(k, "")) for k in columns] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A4B8F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
